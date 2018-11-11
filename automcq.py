#!/usr/bin/python3
    
import sys, platform, random, os.path

#Get OS
host = platform.system()

#Flags
blankmode = False
verbose = False
go = True
gui = False
inputfile = None
outputfile = None

#Parsing command line args
if len(sys.argv) < 2 and not gui:
    print("Feed me a workbook ! Or type \"" + sys.argv[0] + " help\" to get help.")
    go = False

possibleArgs = [ "--blank", "--verbose", "--output", "help", "--gui" ]

for i in range(len(sys.argv)):
    if sys.argv[i][-4:] == "xlsx" and (sys.argv[i-1] != "-o" or sys.argv[i-1] != "--output"):
        if not inputfile:
            inputfile = sys.argv[i]
    elif sys.argv[i] == "--blank":
        blankmode = True
    elif sys.argv[i] == "--verbose":
        verbose = True
    elif sys.argv[i] == "--gui":
        gui = True
    elif sys.argv[i] == "-o" or sys.argv[i] == "--output":
        if i+1 < len(sys.argv):
            if sys.argv[i+1] not in possibleArgs and sys.argv[i+1][0] != '-':
                outputfile = sys.argv[i+1]
            else:
                if host == "Linux":
                    print("\033[91mInvalid output file parameter\033[0m")
                else:
                    print("Invalid output file parameter")

                ok = "z"

                while ok not in "yYnN ":
                    print("OK to write to input file ? [y/N] ", sep="", end="")
                    ok = input()

                    if ok == 'y' or ok == 'Y':
                        go = True
                        outputfile = None
                    elif ok == 'n' or ok == 'N' or ok == '':
                        go = False
        else:
            if host == "Linux":
                print("\033[91mNo output file specified\033[0m")
            else:
                print("No output file specified")

            ok = "z"

            while ok not in "yYnN ":
                print("OK to write to input file ? [y/N] ", sep="", end="")
                ok = input()

                if ok == 'y' or ok == 'Y':
                    go = True
                    outputfile = None
                elif ok == 'n' or ok == 'N' or ok == '':
                    go = False
    elif sys.argv[i][0] == '-' and sys.argv[i][1] != '-':
        if "b" in sys.argv[i]:
            blankmode = True
        if "v" in sys.argv[i]:
            verbose = True
        if "g" in sys.argv[i]:
            gui = True
    elif sys.argv[i] == "help":
        print("AutoMCQ v1.08 by Bad64")
        print("Usage: automcq [switches] [xlsx file]")
        print("    -b or --blank: Only fills in blank cells (do not overwrite filled cells)")
        print("    -v or --verbose: Prints everything to the console")
        print("    -g or --gui: Launches the program in GUI mode")
        print("    -o <file> or --output <file>: Outputs the new filled workbook to file")
        print("        (Note: \"-o\" switch has to be on its own)")
        go = False

#Are we using the GUI ?
if gui:
    try:
        from tkinter import Tk, messagebox
        from tkinter.filedialog import askopenfilename, asksaveasfilename
    except ImportError:
        if host == "Linux":
            print("\033[91mERROR: \033[0mThis program cannot function without tkinter.")
        else:
            print("ERROR: This program cannot function without tkinter.")
        exit(1)
        
    Tk().withdraw()

#Do we even have openpyxl ?
try:
    import openpyxl
except ImportError:
    if gui:
        messagebox.showerror("Error", "This program cannot function without openpyxl.")
    else:
        if host == "Linux":
            print("\033[92mERROR: \033[0mThis program cannot function without openpyxl.")
        else:
            print("ERROR: This program cannot function without openpyxl.")
    exit(1)

#Check if file exists
if not gui:
    if inputfile is not None:
        if not os.path.isfile(inputfile):
            inputfile = None
        else:
            if verbose and go:
                print("Operating on file", inputfile, ":")
else:
    inputfile = askopenfilename(title = "Open", filetypes = (("Excel file","*.xlsx"),("all files","*.*")))

    if not inputfile:
        go = False

#Validating output file
if not gui:
    if outputfile and outputfile[-5:] != ".xlsx":
        outputfile += ".xlsx"
else:
    outputfile = asksaveasfilename(defaultextension = ".xlsx", initialfile = inputfile, title = "Save as", filetypes = (("Excel file","*.xlsx"),("all files","*.*")))

#Main logic
if outputfile is not None and outputfile != inputfile:
    print("Writing to file", outputfile)

if gui and outputfile == inputfile:
    result = messagebox.askquestion("Warning", "Overwrite cells that are already full ?", icon = "warning")

    if (result == "yes"):
        blankmode = False
    else:
        blankmode = True
    
if go and inputfile:
    wb = openpyxl.load_workbook(filename = inputfile)

    sheets = wb.sheetnames
    ws = wb[sheets[0]]

    random.SystemRandom()

    possibleValues = ['A', 'B', 'C', 'D']
    cells = [ 'B2', 'B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B9', 'B10', 'B11', 'B12', 'B13', 'B14', 'B15', 'B16', 'B17', 'B18', 'B19', 'B20', 'B21', 'B22', 'B23', 'B24', 'B25', 'B26', 'B27', 'B28', 'B29', 'B30', 'B31', 'B32', 'B33', 'B34', 'B35', 'B36', 'B37', 'B38', 'B39', 'B40', 'B41' ]
    
    for i in range(len(cells)):
        if blankmode:
            if not ws[cells[i]].value:
                if not verbose:
                    ws[cells[i]] = possibleValues[random.randint(0, 3)]
                else:
                    char = possibleValues[random.randint(0, 3)]
                    if host == "Linux":
                        print("\033[92mAnswering", char, "to question", i+1)
                    else:
                        print("Answering", char, "to question", i+1)
                    ws[cells[i]] = char
            else:
                if verbose:
                    if host == "Linux":
                        print("\033[93mSkipping question", i+1)
                    else:
                        print("Skipping question", i+1)
        else:
            if not verbose:
                ws[cells[i]] = possibleValues[random.randint(0, 3)]
            else:
                char = possibleValues[random.randint(0, 3)]
                if host == "Linux":
                    print("\033[92mAnswering", char, "to question", i+1)
                else:
                    print("Answering", char, "to cell", i+1)
                ws[cells[i]] = char

    if not gui:
        if host == "Linux":
            print("\033[0mDonezo. Now get out of here !")
        else:
            print("Donezo. Now get out of here !")
    else:
        messagebox.showinfo("Success", "Donezo. Now get out of here !")

    if outputfile:
        wb.save(outputfile)
    else:
        wb.save(inputfile)
        
elif go and not inputfile:
    print("You must supply a valid xlsx file !")
elif not go:
    print("", sep="", end="")
